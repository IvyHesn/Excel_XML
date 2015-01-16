from openpyxl import load_workbook
wb = load_workbook(filename='Data_hsn.xlsx', use_iterators=True)
ws = wb.get_sheet_by_name(name='level')  # ws is now an IterableWorksheet

for row in ws.iter_rows():  # it brings a new method: iter_rows()
    for cell in row:
        print (cell.value)
